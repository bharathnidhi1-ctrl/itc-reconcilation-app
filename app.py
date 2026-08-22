"""
app.py
======
GSTR-2B vs. Books — ITC Reconciliation Engine (single-file build)

Everything lives in this one file: the UI-agnostic reconciliation engine,
the vendor-nudge bot, and the Streamlit dashboard on top of them. This is
a straight merge of what used to be three separate modules
(reconciliation_engine.py, nudge_bot.py, app.py) — the logic itself is
unchanged, just co-located so the whole tool ships as a single script.

Run with:
    streamlit run app.py

Need sample data first? Run `python sample_data_generator.py` to create
sample_books.csv / sample_gstr2b.csv, then upload them in the dashboard.
"""

from __future__ import annotations

import io
import re
import logging
import uuid
from collections import defaultdict, deque
from datetime import date
from decimal import Decimal
from typing import Dict, List, Tuple, Optional, Union

import pandas as pd
import streamlit as st
import plotly.express as px
from thefuzz import fuzz
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from database.db_models import (
    Company,
    Invoice,
    ReconciliationRun,
    ReconciliationResult,
    create_database,
)
from ml_models.vendor_risk import VendorRiskModel
from reports.excel_report import generate_excel_report
from reports.pdf_report import generate_pdf_report

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("itc_reconciliation")


# ==========================================================================
# SECTION 1 — Reconciliation Engine (formerly reconciliation_engine.py)
# ==========================================================================
#
# Pipeline
# --------
# 1. Ingestion & Cleaning   -> ITCReconciliationEngine._prepare()
# 2. Tier 1 - Exact Match   -> ITCReconciliationEngine._match_exact_and_tolerance()
# 3. Tier 2 - Tolerance Match (± tolerance, resolves ERP rounding)  -> same method
# 4. Tier 3 - Fuzzy Match (Levenshtein via TheFuzz, same GSTIN)     -> _match_fuzzy()
# 5. Classification into 4 buckets + Excel export

# Canonical schema every input file must be mapped to before it reaches
# the engine. The Streamlit column-mapping UI is responsible for renaming
# raw headers (e.g. "Supplier GSTIN", "Bill No.") to these exact names --
# the engine itself never guesses.
REQUIRED_COLUMNS: List[str] = [
    "GSTIN",
    "Vendor Name",
    "Invoice Number",
    "Invoice Date",
    "Taxable Value",
    "CGST",
    "SGST",
    "IGST",
]

NUMERIC_COLUMNS: List[str] = ["Taxable Value", "CGST", "SGST", "IGST"]

# Internal helper columns created during cleaning. Stripped out of any
# dataframe before it is shown to the user / exported.
_INTERNAL_COLUMNS = ["GSTIN_clean", "Invoice_clean", "_key"]


# --------------------------------------------------------------------------
# Custom Exceptions
# --------------------------------------------------------------------------

class ReconciliationError(Exception):
    """Base class for any business-rule or data-quality failure."""


class SchemaValidationError(ReconciliationError):
    """Raised when a required column is missing from an input dataset."""


class EmptyDatasetError(ReconciliationError):
    """Raised when an uploaded dataset has zero data rows."""


# --------------------------------------------------------------------------
# File I/O helper (shared by the engine and the Streamlit app)
# --------------------------------------------------------------------------

def read_input_file(file: Union[str, "io.BytesIO"], filename: str) -> pd.DataFrame:
    """
    Reads a CSV or Excel file into a DataFrame.

    `file` may be a filesystem path OR a file-like object (e.g. a Streamlit
    UploadedFile / BytesIO buffer). Every column is read in as a string so
    that leading zeros, mixed date formats, etc. survive untouched until
    the engine's cleaning routines process them explicitly. Letting pandas
    auto-infer dtypes here is what silently corrupts GSTINs and invoice
    numbers in most naive reconciliation scripts.
    """
    name = (filename or "").lower()
    try:
        if name.endswith(".csv"):
            return pd.read_csv(file, dtype=str, keep_default_na=True)
        elif name.endswith((".xlsx", ".xls")):
            return pd.read_excel(file, dtype=str)
        else:
            raise ReconciliationError(
                f"Unsupported file format for '{filename}'. Please upload a .csv or .xlsx file."
            )
    except ReconciliationError:
        raise
    except Exception as exc:  # noqa: BLE001 - surface any parser error cleanly
        raise ReconciliationError(f"Failed to read '{filename}': {exc}") from exc


# --------------------------------------------------------------------------
# Main Engine
# --------------------------------------------------------------------------

class ITCReconciliationEngine:
    """
    Configurable 3-tier ITC reconciliation engine.

    Parameters
    ----------
    tolerance : float
        Tier 2 tax-amount tolerance in rupees (default ₹1.00) to absorb
        ERP-side rounding differences.
    fuzzy_threshold : int
        Minimum TheFuzz similarity score (0-100) required for a Tier 3
        invoice-number match under the same GSTIN (default 90).
    """

    def __init__(self, tolerance: float = 1.00, fuzzy_threshold: int = 90):
        if tolerance < 0:
            raise ValueError("tolerance must be >= 0")
        if not (0 <= fuzzy_threshold <= 100):
            raise ValueError("fuzzy_threshold must be between 0 and 100")

        self.tolerance = tolerance
        self.fuzzy_threshold = fuzzy_threshold
        self.warnings: List[str] = []

    # ---------------------------------------------------------------- #
    # Public API
    # ---------------------------------------------------------------- #

    def get_warnings(self) -> List[str]:
        """Data-quality warnings collected during the last reconcile() call."""
        return self.warnings

    def reconcile(self, books_raw: pd.DataFrame, gstr2b_raw: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """
        Runs the full 4-step pipeline and returns a dict of dataframes:
            'ready_to_claim', 'value_mismatches',
            'missing_in_gstr2b', 'missing_in_books', 'summary'
        """
        self.warnings = []

        books = self._prepare(books_raw, "Books (Purchase Register)")
        gstr2b = self._prepare(gstr2b_raw, "GSTR-2B (Portal)")

        # --- Tier 1 (exact) + Tier 2 (tolerance) -----------------------
        ready_1, mismatch_1, mb_idx, mg_idx = self._match_exact_and_tolerance(books, gstr2b)

        unmatched_books = books.drop(index=mb_idx)
        unmatched_gstr2b = gstr2b.drop(index=mg_idx)

        # --- Tier 3 (fuzzy) --------------------------------------------
        ready_3, mismatch_3, fb_idx, fg_idx = self._match_fuzzy(unmatched_books, unmatched_gstr2b)

        missing_in_gstr2b = unmatched_books.drop(index=fb_idx)
        missing_in_books = unmatched_gstr2b.drop(index=fg_idx)

        result: Dict[str, pd.DataFrame] = {
            "ready_to_claim": pd.DataFrame(ready_1 + ready_3),
            "value_mismatches": pd.DataFrame(mismatch_1 + mismatch_3),
            "missing_in_gstr2b": self._finalize_missing(missing_in_gstr2b, note="Vendor Default Risk - Sec 16(2)(aa)"),
            "missing_in_books": self._finalize_missing(missing_in_books, note="Unrecorded Purchase"),
        }
        result["summary"] = self._build_summary(books, gstr2b, result)
        return result

    def export_to_excel(self, results: Dict[str, pd.DataFrame], output: Union[str, "io.BytesIO"]) -> Union[str, "io.BytesIO"]:
        """
        Writes the 4 categorized tables + summary into a single, styled,
        multi-sheet .xlsx workbook using OpenPyXL. `output` may be a file
        path or an in-memory BytesIO buffer (used by the Streamlit app).
        """
        sheet_order = [
            ("summary", "Summary"),
            ("ready_to_claim", "Ready to Claim"),
            ("value_mismatches", "Value Mismatches"),
            ("missing_in_gstr2b", "Missing in GSTR-2B"),
            ("missing_in_books", "Missing in Books"),
        ]

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for key, sheet_name in sheet_order:
                df = results.get(key)
                if df is None or df.empty:
                    df = pd.DataFrame({"Info": [f"No records found in '{sheet_name}'."]})
                # Excel sheet names have a 31-char limit
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

            self._style_workbook(writer.book)

        return output

    # ---------------------------------------------------------------- #
    # Step 1: Ingestion & Cleaning
    # ---------------------------------------------------------------- #

    def _prepare(self, df: pd.DataFrame, label: str) -> pd.DataFrame:
        """Validates schema, then cleans GSTIN / invoice / date / numeric fields."""
        if df is None or len(df) == 0:
            raise EmptyDatasetError(
                f"{label} dataset is empty. Please upload a file that contains data rows."
            )

        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]  # normalize header whitespace

        missing_cols = [c for c in REQUIRED_COLUMNS if c not in df.columns]
        if missing_cols:
            raise SchemaValidationError(
                f"{label} is missing required column(s): {', '.join(missing_cols)}. "
                f"Columns found: {', '.join(df.columns)}"
            )

        df = df.reset_index(drop=True)

        # --- GSTIN cleaning: strip whitespace, uppercase -------------
        df["GSTIN_clean"] = df["GSTIN"].apply(self._clean_gstin)
        bad_gstin = (df["GSTIN_clean"].str.len() != 15).sum()
        if bad_gstin:
            self.warnings.append(
                f"{label}: {bad_gstin} row(s) have a GSTIN that is not the standard 15 characters."
            )

        # --- Invoice Number cleaning: strip specials + leading zeros --
        df["Invoice_clean"] = df["Invoice Number"].apply(self._clean_invoice_number)

        # --- Date standardization to YYYY-MM-DD ------------------------
        parsed_dates = self._parse_dates_robust(df["Invoice Date"])
        n_bad_dates = int(parsed_dates.isna().sum())
        if n_bad_dates:
            self.warnings.append(f"{label}: {n_bad_dates} row(s) had an unparseable Invoice Date.")
        df["Invoice Date"] = parsed_dates.dt.strftime("%Y-%m-%d")

        # --- Numeric normalization (2 decimal places) -------------------
        for col in NUMERIC_COLUMNS:
            coerced = pd.to_numeric(df[col], errors="coerce")
            n_bad = int(coerced.isna().sum())
            if n_bad:
                self.warnings.append(
                    f"{label}: {n_bad} row(s) had a non-numeric value in '{col}' — treated as 0.00."
                )
            df[col] = coerced.fillna(0.0).round(2)

        df["Vendor Name"] = df["Vendor Name"].fillna("Unknown Vendor").astype(str).str.strip()
        df["Vendor Name"] = df["Vendor Name"].replace("", "Unknown Vendor")

        # composite key used for Tier 1 / Tier 2 lookups
        df["_key"] = df["GSTIN_clean"] + "|" + df["Invoice_clean"]

        return df

    @staticmethod
    def _clean_gstin(value) -> str:
        if pd.isna(value):
            return ""
        return re.sub(r"\s+", "", str(value)).strip().upper()

    @staticmethod
    def _parse_dates_robust(series: pd.Series) -> pd.Series:
        """
        Parses a mixed-format date column into proper datetimes.

        IMPORTANT: pandas/dateutil's `dayfirst=True` flag will incorrectly
        swap the month/day of an already-unambiguous ISO date like
        "2025-06-04" (turning it into 2025-04-06) because dateutil applies
        dayfirst to ANY ambiguous two-digit pair it finds, even after a
        clear 4-digit year. A single `pd.to_datetime(..., dayfirst=True)`
        call is therefore NOT safe for the mixed CSV/Excel exports GST
        data typically arrives in (portal exports are often ISO, while
        ERP/Tally exports are typically DD-MM-YYYY or DD/MM/YYYY).

        This method tries formats in order of increasing ambiguity,
        locking in successful matches before ever falling back to a
        dayfirst-based free-form parse:
            1. Strict ISO           (YYYY-MM-DD)
            2. Indian numeric       (DD-MM-YYYY)
            3. Indian numeric       (DD/MM/YYYY)
            4. Free-form fallback   (dateutil, dayfirst=True)
        """
        s = series.astype(str).str.strip()
        originally_blank = series.isna() | (s == "") | (s.str.lower() == "nan")

        parsed = pd.to_datetime(s, format="%Y-%m-%d", errors="coerce")

        mask = parsed.isna() & ~originally_blank
        if mask.any():
            parsed.loc[mask] = pd.to_datetime(s[mask], format="%d-%m-%Y", errors="coerce")

        mask = parsed.isna() & ~originally_blank
        if mask.any():
            parsed.loc[mask] = pd.to_datetime(s[mask], format="%d/%m/%Y", errors="coerce")

        mask = parsed.isna() & ~originally_blank
        if mask.any():
            parsed.loc[mask] = pd.to_datetime(s[mask], errors="coerce", dayfirst=True)

        return parsed

    @staticmethod
    def _clean_invoice_number(value) -> str:
        """
        Normalizes an invoice number for matching purposes:
          1. Cast to string, strip whitespace
          2. Uppercase
          3. Remove ALL special characters (/, -, _, #, spaces, etc.)
          4. Strip leading zeros (keeps a single '0' if the result is empty)
        e.g. "INV/2025-001" and "inv2025001" and "INV 2025 001" all clean to "INV2025001".
        e.g. "0012345" cleans to "12345".
        """
        if pd.isna(value):
            return ""
        s = str(value).strip().upper()
        s = re.sub(r"[^A-Z0-9]", "", s)
        s = s.lstrip("0")
        return s if s else "0"

    # ---------------------------------------------------------------- #
    # Step 2: Tier 1 (Exact) + Tier 2 (Tolerance) Matching
    # ---------------------------------------------------------------- #

    def _match_exact_and_tolerance(
        self, books: pd.DataFrame, gstr2b: pd.DataFrame
    ) -> Tuple[List[dict], List[dict], List[int], List[int]]:
        """
        One-to-one matching on the composite key (GSTIN + cleaned invoice
        number). Uses a deque per key so duplicate invoice numbers under
        the same GSTIN are paired off correctly instead of all colliding
        on a single match. For each matched pair, tax amounts are compared
        to classify the pair as Tier 1 (exact), Tier 2 (within tolerance),
        or a Value Mismatch (same invoice, but tax differs beyond tolerance).
        """
        gstr2b_lookup: Dict[str, deque] = defaultdict(deque)
        for idx, row in gstr2b.iterrows():
            gstr2b_lookup[row["_key"]].append(idx)

        ready_rows, mismatch_rows = [], []
        matched_book_idx, matched_g_idx = [], []

        for b_idx, b_row in books.iterrows():
            bucket = gstr2b_lookup.get(b_row["_key"])
            if not bucket:
                continue
            g_idx = bucket.popleft()
            g_row = gstr2b.loc[g_idx]

            tier, max_diff = self._classify_pair(b_row, g_row)
            result_row = self._build_result_row(b_row, g_row, tier, max_diff)

            matched_book_idx.append(b_idx)
            matched_g_idx.append(g_idx)

            if tier == "Value Mismatch":
                mismatch_rows.append(result_row)
            else:
                ready_rows.append(result_row)

        return ready_rows, mismatch_rows, matched_book_idx, matched_g_idx

    def _classify_pair(self, b_row: pd.Series, g_row: pd.Series) -> Tuple[str, float]:
        """Compares tax amounts of an already invoice-number-matched pair."""
        diffs = [
            abs(b_row["Taxable Value"] - g_row["Taxable Value"]),
            abs(b_row["CGST"] - g_row["CGST"]),
            abs(b_row["SGST"] - g_row["SGST"]),
            abs(b_row["IGST"] - g_row["IGST"]),
        ]
        max_diff = max(diffs)

        if max_diff < 0.005:  # float-safe zero (handles 2-decimal rounding noise)
            return "Tier 1 - Exact Match", max_diff
        elif max_diff <= self.tolerance:
            return "Tier 2 - Tolerance Match", max_diff
        else:
            return "Value Mismatch", max_diff

    # ---------------------------------------------------------------- #
    # Step 3: Tier 3 - Fuzzy Matching
    # ---------------------------------------------------------------- #

    def _match_fuzzy(
        self, unmatched_books: pd.DataFrame, unmatched_gstr2b: pd.DataFrame
    ) -> Tuple[List[dict], List[dict], List[int], List[int]]:
        """
        For rows left unmatched after Tier 1/2, applies Levenshtein-based
        fuzzy matching (TheFuzz `ratio`) between invoice numbers, but only
        within the SAME GSTIN (comparing invoice numbers across unrelated
        vendors would be meaningless and dangerous for a tax reconciliation).

        Candidate pairs scoring >= fuzzy_threshold are greedily assigned
        highest-score-first, one-to-one, so no invoice is matched twice.
        Amount comparison then still applies (reusing _classify_pair) so a
        fuzzy invoice-number match with a large tax discrepancy correctly
        lands in 'Value Mismatches' rather than being silently accepted.
        """
        ready_rows, mismatch_rows = [], []
        matched_book_idx, matched_g_idx = [], []

        if unmatched_books.empty or unmatched_gstr2b.empty:
            return ready_rows, mismatch_rows, matched_book_idx, matched_g_idx

        for gstin, b_group in unmatched_books.groupby("GSTIN_clean"):
            if not gstin:
                continue  # never fuzzy-match rows with a blank/invalid GSTIN
            g_group = unmatched_gstr2b[unmatched_gstr2b["GSTIN_clean"] == gstin]
            if g_group.empty:
                continue

            candidate_pairs = []
            for b_idx, b_row in b_group.iterrows():
                for g_idx, g_row in g_group.iterrows():
                    score = fuzz.ratio(b_row["Invoice_clean"], g_row["Invoice_clean"])
                    if score >= self.fuzzy_threshold:
                        candidate_pairs.append((score, b_idx, g_idx))

            # Greedy assignment: highest similarity first, one-to-one
            candidate_pairs.sort(key=lambda x: -x[0])
            used_b, used_g = set(), set()

            for score, b_idx, g_idx in candidate_pairs:
                if b_idx in used_b or g_idx in used_g:
                    continue
                used_b.add(b_idx)
                used_g.add(g_idx)

                b_row = unmatched_books.loc[b_idx]
                g_row = unmatched_gstr2b.loc[g_idx]
                base_tier, max_diff = self._classify_pair(b_row, g_row)

                if base_tier == "Value Mismatch":
                    tier_label = f"Value Mismatch (Fuzzy Invoice Match, {score}% similarity)"
                else:
                    tier_label = f"Tier 3 - Fuzzy Match ({score}% similarity)"

                result_row = self._build_result_row(b_row, g_row, tier_label, max_diff)
                matched_book_idx.append(b_idx)
                matched_g_idx.append(g_idx)

                if base_tier == "Value Mismatch":
                    mismatch_rows.append(result_row)
                else:
                    ready_rows.append(result_row)

        return ready_rows, mismatch_rows, matched_book_idx, matched_g_idx

    # ---------------------------------------------------------------- #
    # Step 4: Row builders, missing-bucket formatting, summary, export
    # ---------------------------------------------------------------- #

    @staticmethod
    def _build_result_row(b_row: pd.Series, g_row: pd.Series, tier: str, max_diff: float) -> dict:
        return {
            "GSTIN": b_row["GSTIN_clean"],
            "Vendor Name (Books)": b_row.get("Vendor Name", ""),
            "Vendor Name (GSTR-2B)": g_row.get("Vendor Name", ""),
            "Invoice Number (Books)": b_row.get("Invoice Number", ""),
            "Invoice Number (GSTR-2B)": g_row.get("Invoice Number", ""),
            "Invoice Date (Books)": b_row.get("Invoice Date", ""),
            "Invoice Date (GSTR-2B)": g_row.get("Invoice Date", ""),
            "Taxable Value (Books)": round(float(b_row["Taxable Value"]), 2),
            "Taxable Value (GSTR-2B)": round(float(g_row["Taxable Value"]), 2),
            "CGST (Books)": round(float(b_row["CGST"]), 2),
            "CGST (GSTR-2B)": round(float(g_row["CGST"]), 2),
            "SGST (Books)": round(float(b_row["SGST"]), 2),
            "SGST (GSTR-2B)": round(float(g_row["SGST"]), 2),
            "IGST (Books)": round(float(b_row["IGST"]), 2),
            "IGST (GSTR-2B)": round(float(g_row["IGST"]), 2),
            "Max Tax Diff (₹)": round(float(max_diff), 2),
            "Match Tier": tier,
        }

    @staticmethod
    def _finalize_missing(df: pd.DataFrame, note: str) -> pd.DataFrame:
        """Drops internal helper columns and tags the risk/reason note."""
        if df.empty:
            cols = [c for c in REQUIRED_COLUMNS] + ["Remarks"]
            return pd.DataFrame(columns=cols)
        out = df.drop(columns=_INTERNAL_COLUMNS, errors="ignore").copy()
        out["Remarks"] = note
        return out.reset_index(drop=True)

    def _build_summary(
        self, books: pd.DataFrame, gstr2b: pd.DataFrame, result: Dict[str, pd.DataFrame]
    ) -> pd.DataFrame:
        ready = result["ready_to_claim"]
        mismatches = result["value_mismatches"]
        missing_2b = result["missing_in_gstr2b"]
        missing_books = result["missing_in_books"]

        matched_itc = 0.0
        if not ready.empty:
            matched_itc = round(
                ready[["CGST (Books)", "SGST (Books)", "IGST (Books)"]].sum().sum(), 2
            )

        blocked_itc = 0.0
        if not missing_2b.empty:
            blocked_itc = round(missing_2b[["CGST", "SGST", "IGST"]].sum().sum(), 2)

        unrecorded_itc = 0.0
        if not missing_books.empty:
            unrecorded_itc = round(missing_books[["CGST", "SGST", "IGST"]].sum().sum(), 2)

        mismatch_diff_total = 0.0
        if not mismatches.empty:
            mismatch_diff_total = round(mismatches["Max Tax Diff (₹)"].sum(), 2)

        vendor_risk_count = 0
        if not missing_2b.empty and "Vendor Name" in missing_2b.columns:
            vendor_risk_count = int(missing_2b["Vendor Name"].nunique())

        data = {
            "Metric": [
                "Total Invoices in Books",
                "Total Invoices in GSTR-2B",
                "Ready to Claim (Matched Count)",
                "Value Mismatch Count",
                "Missing in GSTR-2B (Count)",
                "Missing in Books (Count)",
                "Matched ITC Value (Rs)",
                "Blocked ITC Value (Rs) - Sec 16(2)(aa) Risk",
                "Unrecorded ITC in Books (Rs)",
                "Total Tax Diff in Mismatches (Rs)",
                "Vendors at Risk (Missing in 2B)",
            ],
            "Value": [
                len(books),
                len(gstr2b),
                len(ready),
                len(mismatches),
                len(missing_2b),
                len(missing_books),
                matched_itc,
                blocked_itc,
                unrecorded_itc,
                mismatch_diff_total,
                vendor_risk_count,
            ],
        }
        return pd.DataFrame(data)

    @staticmethod
    def _style_workbook(workbook) -> None:
        """Applies light corporate styling: bold white-on-navy headers,
        frozen header row, and auto-sized columns."""
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.freeze_panes = "A2"

            for col_cells in sheet.columns:
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col_cells
                )
                col_letter = get_column_letter(col_cells[0].column)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)


# ==========================================================================
# SECTION 2 — Vendor Nudge Bot (formerly nudge_bot.py)
# ==========================================================================
#
# Takes the 'Missing in GSTR-2B' dataframe produced by
# ITCReconciliationEngine, groups blocked ITC by (Vendor Name, GSTIN), and
# drafts compliance-ready Email + WhatsApp notices referencing Section
# 16(2)(aa) of the CGST Act, 2017, asking the supplier to file/amend their
# GSTR-1 before a statutory cutoff date.

REQUIRED_NUDGE_COLUMNS = {"Vendor Name", "GSTIN", "Taxable Value", "CGST", "SGST", "IGST", "Invoice Number", "Invoice Date"}


def generate_vendor_nudges(
    missing_in_gstr2b_df: pd.DataFrame,
    cutoff_date: Optional[str] = None,
    company_name: str = "[Your Company Name]",
    contact_person: str = "[GST / Accounts Team]",
    max_invoices_listed: int = 10,
) -> List[Dict]:
    """
    Builds one nudge record per defaulting vendor.

    Parameters
    ----------
    missing_in_gstr2b_df : DataFrame
        Output of results['missing_in_gstr2b'] from the reconciliation engine.
    cutoff_date : str, optional
        Statutory/internal cutoff date to display in the drafts (e.g. "11-Sep-2026").
        Defaults to the 11th of next month if not supplied.
    company_name, contact_person : str
        Used to personalize the drafts.
    max_invoices_listed : int
        Caps how many individual invoices are itemized in the email body.

    Returns
    -------
    List[dict] — one entry per vendor with keys:
        'Vendor Name', 'GSTIN', 'Invoice Count', 'Blocked Tax (₹)',
        'Email Draft', 'WhatsApp Draft'
    """
    if missing_in_gstr2b_df is None or missing_in_gstr2b_df.empty:
        return []

    missing_cols = REQUIRED_NUDGE_COLUMNS - set(missing_in_gstr2b_df.columns)
    if missing_cols:
        raise ValueError(
            f"Cannot generate vendor nudges — missing required column(s): {', '.join(sorted(missing_cols))}"
        )

    if not cutoff_date:
        cutoff_date = _default_cutoff_date()

    df = missing_in_gstr2b_df.copy()
    df["Total Tax (Rs)"] = df["CGST"] + df["SGST"] + df["IGST"]

    grouped = (
        df.groupby(["Vendor Name", "GSTIN"], dropna=False)
        .agg(
            Invoice_Count=("Invoice Number", "count"),
            Blocked_Taxable_Value=("Taxable Value", "sum"),
            Blocked_Tax=("Total Tax (Rs)", "sum"),
        )
        .reset_index()
        .sort_values("Blocked_Tax", ascending=False)
    )

    drafts: List[Dict] = []
    for _, row in grouped.iterrows():
        vendor = row["Vendor Name"]
        gstin = row["GSTIN"]
        invoice_count = int(row["Invoice_Count"])
        blocked_tax = round(float(row["Blocked_Tax"]), 2)
        blocked_taxable = round(float(row["Blocked_Taxable_Value"]), 2)

        vendor_invoices = df[(df["Vendor Name"] == vendor) & (df["GSTIN"] == gstin)]

        email_draft = _build_email_draft(
            vendor=vendor,
            gstin=gstin,
            invoice_count=invoice_count,
            blocked_tax=blocked_tax,
            blocked_taxable=blocked_taxable,
            cutoff_date=cutoff_date,
            company_name=company_name,
            contact_person=contact_person,
            vendor_invoices=vendor_invoices,
            max_invoices_listed=max_invoices_listed,
        )
        whatsapp_draft = _build_whatsapp_draft(
            vendor=vendor,
            gstin=gstin,
            invoice_count=invoice_count,
            blocked_tax=blocked_tax,
            cutoff_date=cutoff_date,
            company_name=company_name,
        )

        drafts.append(
            {
                "Vendor Name": vendor,
                "GSTIN": gstin,
                "Invoice Count": invoice_count,
                "Blocked Tax (Rs)": blocked_tax,
                "Email Draft": email_draft,
                "WhatsApp Draft": whatsapp_draft,
            }
        )

    return drafts


def _default_cutoff_date() -> str:
    """
    Defaults to the 11th of next month (the standard GSTR-1 due date rhythm
    for most taxpayers). Wire this to your actual statutory due-date
    calendar / QRMP scheme dates in a production deployment.
    """
    today = date.today()
    year = today.year + (1 if today.month == 12 else 0)
    month = 1 if today.month == 12 else today.month + 1
    return date(year, month, 11).strftime("%d-%b-%Y")


def _build_email_draft(
    vendor: str,
    gstin: str,
    invoice_count: int,
    blocked_tax: float,
    blocked_taxable: float,
    cutoff_date: str,
    company_name: str,
    contact_person: str,
    vendor_invoices: pd.DataFrame,
    max_invoices_listed: int,
) -> str:
    listed = vendor_invoices.head(max_invoices_listed)
    invoice_lines = "\n".join(
        f"   - Invoice {r['Invoice Number']} dated {r['Invoice Date']} "
        f"(Taxable Value Rs {r['Taxable Value']:.2f}, Tax Rs {(r['CGST'] + r['SGST'] + r['IGST']):.2f})"
        for _, r in listed.iterrows()
    )
    remainder = len(vendor_invoices) - len(listed)
    more_note = f"\n   ...and {remainder} more invoice(s) — see attached reconciliation report." if remainder > 0 else ""

    return f"""Subject: Action Required — Invoices Missing in GSTR-2B (Cutoff: {cutoff_date})

Dear {vendor} Team,

This is an automated ITC compliance notice from {company_name}.

During our routine reconciliation of the Purchase Register against the
GSTR-2B statement auto-populated on the GST Portal, we identified
{invoice_count} invoice(s) totaling Rs {blocked_taxable:.2f} in taxable value
(Rs {blocked_tax:.2f} in tax) that are recorded in our books under GSTIN
{gstin} but are currently NOT reflected in our GSTR-2B.

As per Section 16(2)(aa) of the CGST Act, 2017, we are entitled to avail
Input Tax Credit only on invoices that have been furnished by the supplier
in their GSTR-1/IFF and which consequently appear in our GSTR-2B. Until
these invoices are reported, we are unable to claim the related ITC.

Affected Invoices:
{invoice_lines}{more_note}

We request you to:
   1. Verify whether the above invoices have been reported in your GSTR-1/IFF.
   2. File or amend your GSTR-1 to include any missing invoices at the earliest.
   3. Confirm your expected filing date to our team.

Kindly complete the above before {cutoff_date} to avoid disruption to our
ongoing business relationship and to ensure timely ITC availment.

For any queries, please reach out to {contact_person}.

Regards,
{contact_person}
{company_name}

(This is a system-generated compliance communication from the ITC
Reconciliation Engine. Please do not ignore.)
"""


def _build_whatsapp_draft(
    vendor: str, gstin: str, invoice_count: int, blocked_tax: float, cutoff_date: str, company_name: str
) -> str:
    return (
        f"*ITC Reconciliation Alert — {company_name}*\n\n"
        f"Hi {vendor} Team,\n"
        f"We found *{invoice_count} invoice(s)* (GSTIN: {gstin}) worth "
        f"*Rs {blocked_tax:.2f}* tax in our books that are *missing from GSTR-2B*.\n\n"
        f"As per Sec 16(2)(aa) of the CGST Act, we cannot claim this ITC until "
        f"you file/amend your GSTR-1. Kindly upload these invoices before "
        f"*{cutoff_date}*.\n\n"
        f"Please confirm once done. Thank you!"
    )


def _as_date(value) -> Optional[date]:
    parsed = pd.to_datetime(value, errors="coerce")
    return None if pd.isna(parsed) else parsed.date()


def save_reconciliation_run(
    books_df: pd.DataFrame,
    gstr2b_df: pd.DataFrame,
    results: Dict[str, pd.DataFrame],
    company_name: str,
) -> Tuple[str, pd.DataFrame]:
    """Persist one reconciliation run and its vendor risk assessment."""
    engine = create_database()
    run_id = uuid.uuid4().hex

    with Session(engine) as session:
        company = session.scalar(select(Company).where(Company.legal_name == company_name))
        if company is None:
            company = Company(legal_name=company_name)
            session.add(company)
            session.flush()

        session.add(
            ReconciliationRun(
                id=run_id,
                company_id=company.id,
                total_books=len(books_df),
                total_gstr2b=len(gstr2b_df),
                matched_count=len(results["ready_to_claim"]),
                mismatch_count=len(results["value_mismatches"]),
                missing_in_gstr2b_count=len(results["missing_in_gstr2b"]),
                missing_in_books_count=len(results["missing_in_books"]),
            )
        )

        invoice_ids = {}
        for source, frame in (("books", books_df), ("gstr2b", gstr2b_df)):
            for _, row in frame.iterrows():
                gstin = str(row["GSTIN"]).strip().upper()
                invoice_number = str(row["Invoice Number"]).strip()
                key = (source, gstin, invoice_number)
                invoice = session.scalar(
                    select(Invoice).where(
                        Invoice.company_id == company.id,
                        Invoice.source == source,
                        Invoice.gstin == gstin,
                        Invoice.invoice_number == invoice_number,
                    )
                )
                if invoice is None:
                    invoice = Invoice(
                        company_id=company.id,
                        source=source,
                        gstin=gstin,
                        vendor_name=str(row["Vendor Name"]),
                        invoice_number=invoice_number,
                        invoice_date=_as_date(row["Invoice Date"]),
                        taxable_value=Decimal(str(row["Taxable Value"])),
                        cgst=Decimal(str(row["CGST"])),
                        sgst=Decimal(str(row["SGST"])),
                        igst=Decimal(str(row["IGST"])),
                    )
                    session.add(invoice)
                    session.flush()
                invoice_ids[key] = invoice.id

        category_status = {
            "ready_to_claim": "Ready to Claim",
            "value_mismatches": "Value Mismatch",
            "missing_in_gstr2b": "Missing in GSTR-2B",
            "missing_in_books": "Missing in Books",
        }
        for category, status in category_status.items():
            for _, row in results[category].iterrows():
                gstin = str(row.get("GSTIN", "")).strip().upper()
                books_number = str(row.get("Invoice Number (Books)", "")).strip()
                gstr_number = str(row.get("Invoice Number (GSTR-2B)", "")).strip()
                if category == "missing_in_books":
                    invoice_id = invoice_ids.get(("gstr2b", gstin, gstr_number))
                else:
                    invoice_id = invoice_ids.get(("books", gstin, books_number))
                taxable_difference = abs(
                    float(row.get("Taxable Value (Books)", row.get("Taxable Value", 0)))
                    - float(row.get("Taxable Value (GSTR-2B)", 0))
                )
                session.add(
                    ReconciliationResult(
                        company_id=company.id,
                        invoice_id=invoice_id,
                        run_id=run_id,
                        status=status,
                        match_tier=row.get("Match Tier"),
                        taxable_value_difference=Decimal(str(round(taxable_difference, 2))),
                        tax_difference=Decimal(str(row.get("Max Tax Diff (₹)", 0))),
                        remarks=row.get("Remarks"),
                    )
                )

        session.commit()
        company_id = company.id

    history = pd.concat(
        [results[key] for key in category_status], ignore_index=True, sort=False
    )
    risk_model = VendorRiskModel()
    risk_scores = risk_model.calculate_vendor_risk(history)
    risk_model.save_risk_scores(engine, risk_scores, company_id)
    return run_id, risk_scores


# ==========================================================================
# SECTION 3 — Streamlit Dashboard (formerly app.py)
# ==========================================================================
# Presentation only (upload, column mapping, metrics, tabs, downloads).
# Everything above this line is the reusable, UI-agnostic business logic.

st.set_page_config(
    page_title="ITC Reconciliation Engine | GSTR-2B vs Books",
    page_icon="🧾",
    layout="wide",
)

st.markdown(
    """
    <style>
    [data-testid="stMetric"] {
        background: #f4f7f9;
        border: 1px solid #d9e2e8;
        border-radius: 10px;
        padding: 12px 14px;
    }
    [data-testid="stMetricValue"] { color: #173f5f; }
    div[data-baseweb="tab-list"] { gap: 8px; }
    button[kind="primary"] { background: #1f4e78; border-color: #1f4e78; }
    </style>
    """,
    unsafe_allow_html=True,
)


# --------------------------------------------------------------------------
# Column-mapping helpers
# --------------------------------------------------------------------------

def guess_column_mapping(columns: list, required_fields: list) -> dict:
    """Best-effort auto-mapping of raw file headers to the engine's
    canonical field names, using fuzzy string similarity. The user always
    confirms/overrides this in the UI -- it's a convenience, not a
    silent assumption."""
    mapping = {}
    for field in required_fields:
        best_col, best_score = None, 0
        for col in columns:
            score = fuzz.token_sort_ratio(field.lower(), str(col).lower())
            if score > best_score:
                best_score, best_col = score, col
        mapping[field] = best_col if best_score >= 60 else None
    return mapping


def render_upload_and_mapping(label: str, key_prefix: str):
    """Renders a file uploader + column-mapping UI for one input file.
    Returns a cleaned, canonically-named DataFrame ready for the engine,
    or None if the file isn't uploaded / mapped yet."""
    st.subheader(label)
    file = st.file_uploader(
        f"Upload {label} (.csv or .xlsx)", type=["csv", "xlsx", "xls"], key=f"{key_prefix}_file"
    )
    if not file:
        return None

    try:
        raw_df = read_input_file(file, file.name)
    except ReconciliationError as e:
        st.error(str(e))
        return None

    if raw_df.empty:
        st.error(f"{label} file has no data rows.")
        return None

    st.caption(f"Detected {len(raw_df)} row(s) and {len(raw_df.columns)} column(s).")

    all_present = all(c in raw_df.columns for c in REQUIRED_COLUMNS)
    with st.expander(f"🔧 Map columns for {label}", expanded=not all_present):
        st.caption(
            "We auto-suggest the best matching column for each required field. "
            "Please review and correct any mapping before running reconciliation."
        )
        guessed = guess_column_mapping(list(raw_df.columns), REQUIRED_COLUMNS)
        mapping = {}
        ui_cols = st.columns(2)
        for i, field in enumerate(REQUIRED_COLUMNS):
            options = ["-- Select Column --"] + list(raw_df.columns)
            default_idx = options.index(guessed[field]) if guessed[field] in options else 0
            with ui_cols[i % 2]:
                choice = st.selectbox(field, options, index=default_idx, key=f"{key_prefix}_{field}")
            mapping[field] = None if choice == "-- Select Column --" else choice

    unmapped = [f for f, v in mapping.items() if v is None]
    if unmapped:
        st.warning(f"Please map the following required field(s): {', '.join(unmapped)}")
        return None

    mapped_df = raw_df.rename(columns={v: k for k, v in mapping.items()})[REQUIRED_COLUMNS].copy()
    st.dataframe(mapped_df.head(5), use_container_width=True)
    return mapped_df


# --------------------------------------------------------------------------
# Sidebar - engine & nudge bot settings
# --------------------------------------------------------------------------

st.sidebar.header("⚙️ Matching Engine Settings")
tolerance = st.sidebar.number_input(
    "Tier 2 Tolerance (₹)", min_value=0.0, max_value=100.0, value=1.00, step=0.5,
    help="Maximum rupee difference in any tax field to still be treated as an ERP rounding error.",
)
fuzzy_threshold = st.sidebar.slider(
    "Tier 3 Fuzzy Match Threshold (%)", min_value=70, max_value=100, value=90,
    help="Minimum invoice-number similarity score (Levenshtein-based) required for a Tier 3 match.",
)

st.sidebar.divider()
st.sidebar.header("📨 Nudge Bot Settings")
company_name = st.sidebar.text_input("Your Company Name", "ABC Enterprises Pvt. Ltd.")
contact_person = st.sidebar.text_input("Contact Person / Team", "GST Compliance Team")
cutoff_date_input = st.sidebar.date_input("Vendor Cutoff Date", value=None)

st.sidebar.divider()
st.sidebar.caption(
    "Built with Pandas, NumPy, TheFuzz & Streamlit · 3-Tier Matching: "
    "Exact → Tolerance → Fuzzy (Levenshtein)"
)

# --------------------------------------------------------------------------
# Main layout
# --------------------------------------------------------------------------

st.title("🧾 GSTR-2B vs. Books — ITC Reconciliation Engine")
st.caption(
    "Enterprise-grade Input Tax Credit reconciliation with intelligent 3-tier "
    "matching and an automated vendor-nudge workflow."
)

col1, col2 = st.columns(2)
with col1:
    books_df = render_upload_and_mapping("Purchase Register (Books)", "books")
with col2:
    gstr2b_df = render_upload_and_mapping("GSTR-2B Statement (Portal)", "gstr2b")

st.divider()

run = st.button(
    "🚀 Run Reconciliation",
    type="primary",
    use_container_width=True,
    disabled=not (books_df is not None and gstr2b_df is not None),
)

if run:
    engine = ITCReconciliationEngine(tolerance=tolerance, fuzzy_threshold=fuzzy_threshold)
    try:
        with st.spinner("Cleaning data and running the 3-tier matching engine..."):
            results = engine.reconcile(books_df, gstr2b_df)
            run_id, risk_scores = save_reconciliation_run(
                books_df, gstr2b_df, results, company_name
            )
        st.session_state["results"] = results
        st.session_state["risk_scores"] = risk_scores
        st.session_state["run_id"] = run_id
        st.session_state["warnings"] = engine.get_warnings()
        st.session_state["engine_settings"] = (tolerance, fuzzy_threshold)
        st.success(f"Reconciliation complete and saved to gst_reconciliation.db (run {run_id[:8]}).")
    except ReconciliationError as e:
        st.error(f"Reconciliation failed: {e}")
        st.stop()
    except Exception as e:  # noqa: BLE001 - keep database/model failures visible in the UI
        logger.exception("Failed to persist reconciliation run or calculate vendor risk")
        st.error(f"Reconciliation completed, but saving analytics failed: {e}")
        st.stop()

# --------------------------------------------------------------------------
# Results
# --------------------------------------------------------------------------

if "results" in st.session_state:
    results = st.session_state["results"]
    warnings = st.session_state.get("warnings", [])

    if warnings:
        with st.expander(f"⚠️ {len(warnings)} Data Quality Warning(s) — click to review"):
            for w in warnings:
                st.write(f"- {w}")

    summary = results["summary"].set_index("Metric")["Value"]

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Invoices (Books)", int(summary.get("Total Invoices in Books", 0)))
    m2.metric("Total Invoices (GSTR-2B)", int(summary.get("Total Invoices in GSTR-2B", 0)))
    m3.metric("Matched ITC (₹)", f"{summary.get('Matched ITC Value (Rs)', 0):,.2f}")
    m4.metric(
        "Blocked ITC (₹)",
        f"{summary.get('Blocked ITC Value (Rs) - Sec 16(2)(aa) Risk', 0):,.2f}",
    )
    m5.metric("Vendor Risk Count", int(summary.get("Vendors at Risk (Missing in 2B)", 0)))

    tabs = st.tabs(
        [
            "✅ Ready to Claim",
            "🚫 Missing in GSTR-2B",
            "📥 Missing in Books",
            "⚠️ Value Mismatches",
            "📨 Vendor Nudge Bot",
            "📈 Executive Risk & Analytics",
            "📊 Summary",
        ]
    )

    with tabs[0]:
        st.caption("Fully matched (Tier 1) + resolved within tolerance / fuzzy match (Tier 2 & 3). Eligible for ITC claim.")
        st.dataframe(results["ready_to_claim"], use_container_width=True)

    with tabs[1]:
        st.caption("Present in Books, absent in GSTR-2B — vendor default risk under Section 16(2)(aa) of the CGST Act.")
        st.dataframe(results["missing_in_gstr2b"], use_container_width=True)

    with tabs[2]:
        st.caption("Present in GSTR-2B, absent in Books — possible unrecorded purchase.")
        st.dataframe(results["missing_in_books"], use_container_width=True)

    with tabs[3]:
        st.caption("Invoice exists on both sides, but tax amounts differ beyond the configured tolerance.")
        st.dataframe(results["value_mismatches"], use_container_width=True)

    with tabs[4]:
        st.subheader("Automated Vendor Nudge Drafts")
        missing_2b_raw = results["missing_in_gstr2b"]
        if missing_2b_raw.empty:
            st.info("No defaulting vendors found — every invoice in Books is reflected in GSTR-2B. 🎉")
        else:
            cutoff_str = cutoff_date_input.strftime("%d-%b-%Y") if cutoff_date_input else None
            try:
                drafts = generate_vendor_nudges(
                    missing_2b_raw,
                    cutoff_date=cutoff_str,
                    company_name=company_name,
                    contact_person=contact_person,
                )
            except ValueError as e:
                drafts = []
                st.error(str(e))

            for d in drafts:
                header = f"{d['Vendor Name']} ({d['GSTIN']}) — {d['Invoice Count']} invoice(s), ₹{d['Blocked Tax (Rs)']:,.2f} blocked"
                with st.expander(header):
                    st.text_area("Email Draft", d["Email Draft"], height=340, key=f"email_{d['GSTIN']}")
                    st.text_area("WhatsApp Draft", d["WhatsApp Draft"], height=160, key=f"wa_{d['GSTIN']}")

    with tabs[5]:
        st.subheader("Executive Risk & Analytics Dashboard")
        risk_scores = st.session_state.get("risk_scores", pd.DataFrame())
        match_status = pd.DataFrame(
            {
                "Match Status": [
                    "Ready to Claim",
                    "Value Mismatch",
                    "Missing in GSTR-2B",
                    "Missing in Books",
                ],
                "Invoices": [
                    len(results["ready_to_claim"]),
                    len(results["value_mismatches"]),
                    len(results["missing_in_gstr2b"]),
                    len(results["missing_in_books"]),
                ],
            }
        )
        match_chart = px.pie(
            match_status,
            names="Match Status",
            values="Invoices",
            hole=0.45,
            color="Match Status",
            color_discrete_map={
                "Ready to Claim": "#287d58",
                "Value Mismatch": "#e67e22",
                "Missing in GSTR-2B": "#c0392b",
                "Missing in Books": "#61758a",
            },
            title="Invoice Match Status",
        )

        if risk_scores.empty:
            st.info("No vendor risk scores were produced for this reconciliation run.")
            risk_level_counts = pd.DataFrame(
                {"Risk Category": ["Low", "Medium", "High"], "Vendors": [0, 0, 0]}
            )
        else:
            category_column = "Risk Category" if "Risk Category" in risk_scores else "risk_level"
            risk_level_counts = (
                risk_scores[category_column]
                .rename("Risk Category")
                .value_counts()
                .reindex(["Low", "Medium", "High"], fill_value=0)
                .rename_axis("Risk Category")
                .reset_index(name="Vendors")
            )
            high_risk = int((risk_scores[category_column] == "High").sum())
            total_blocked = float(risk_scores["blocked_itc"].sum())
            r1, r2, r3 = st.columns(3)
            r1.metric("Vendors Assessed", len(risk_scores))
            r2.metric("High Risk Vendors", high_risk)
            r3.metric("Blocked ITC in Risk Model (₹)", f"{total_blocked:,.2f}")

        risk_chart = px.bar(
            risk_level_counts,
            x="Risk Category",
            y="Vendors",
            color="Risk Category",
            text="Vendors",
            color_discrete_map={"High": "#c0392b", "Medium": "#e67e22", "Low": "#287d58"},
            title="Vendor Risk Levels",
        )
        chart_cols = st.columns(2)
        with chart_cols[0]:
            st.plotly_chart(match_chart, use_container_width=True)
        with chart_cols[1]:
            st.plotly_chart(risk_chart, use_container_width=True)
        if not risk_scores.empty:
            st.dataframe(risk_scores, use_container_width=True, hide_index=True)

    with tabs[6]:
        st.dataframe(results["summary"], use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("⬇️ Export Reconciled Report")

    excel_buffer = generate_excel_report(results)
    pdf_buffer = generate_pdf_report(
        summary=results["summary"],
        risk_scores=st.session_state.get("risk_scores"),
    )
    export_cols = st.columns(2)
    with export_cols[0]:
        st.download_button(
            "Download Excel Report (.xlsx)",
            data=excel_buffer.getvalue(),
            file_name=f"ITC_Reconciliation_Report_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with export_cols[1]:
        st.download_button(
            "Download Executive Audit (.pdf)",
            data=pdf_buffer.getvalue(),
            file_name=f"Executive_Risk_Reconciliation_Audit_{date.today().isoformat()}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

else:
    st.info("Upload both files above, confirm the column mapping, then click **Run Reconciliation** to begin.")
    st.caption(
        "Need sample data to try this out? Run `python sample_data_generator.py` "
        "to create `sample_books.csv` and `sample_gstr2b.csv`."
    )
