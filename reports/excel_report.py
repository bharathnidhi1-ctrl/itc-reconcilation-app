"""Excel export helpers for reconciliation and vendor-risk reports."""

from __future__ import annotations

from io import BytesIO
from os import PathLike
from pathlib import Path
from typing import Mapping, Optional, Union

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RESULT_SHEETS = (
    ("summary", "Summary"),
    ("ready_to_claim", "Matched Invoices"),
    ("value_mismatches", "Mismatched Invoices"),
    ("missing_in_gstr2b", "Missing in GSTR-2B"),
    ("missing_in_books", "Missing in Books"),
)

OutputTarget = Union[str, PathLike[str], BytesIO]


def _sheet_frame(data: Optional[pd.DataFrame], sheet_name: str) -> pd.DataFrame:
    if data is None or data.empty:
        return pd.DataFrame({"Info": [f"No records found in '{sheet_name}'."]})
    return data.copy()


def _style_workbook(workbook) -> None:
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    stripe_fill = PatternFill(
        start_color="EAF2F8", end_color="EAF2F8", fill_type="solid"
    )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = stripe_fill
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in cells) + 3
            letter = get_column_letter(cells[0].column)
            sheet.column_dimensions[letter].width = min(max(width, 12), 42)


def generate_excel_report(
    results: Mapping[str, pd.DataFrame],
    output: Optional[OutputTarget] = None,
) -> BytesIO:
    """Export reconciliation dataframes to styled Excel sheets.

    ``output`` may be a filesystem path or an in-memory buffer. The returned
    buffer is always positioned at byte zero for direct Streamlit downloads.
    """
    buffer = output if isinstance(output, BytesIO) else BytesIO()
    buffer.seek(0)
    buffer.truncate(0)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for key, sheet_name in RESULT_SHEETS:
            _sheet_frame(results.get(key), sheet_name).to_excel(
                writer, sheet_name=sheet_name[:31], index=False
            )
        _style_workbook(writer.book)

    buffer.seek(0)
    if isinstance(output, (str, PathLike)):
        with Path(output).open("wb") as file:
            file.write(buffer.getvalue())
        buffer.seek(0)
    return buffer


if __name__ == "__main__":
    generate_excel_report({})
    print("Excel report buffer generated.")
